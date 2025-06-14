import pandas as pd
import glob
from fpdf import FPDF, YPos, XPos
from pathlib import Path

from openpyxl.drawing.xdr import XDRPositiveSize2D

filepaths = glob.glob("invoices/*.xlsx")


for filepath in filepaths:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_nr, invoice_date = filename.split("-")

    pdf.set_font("Times", size=16, style="B")
    pdf.cell(50, 8, text=f"Invoice nr.{invoice_nr}", new_y=YPos.NEXT, new_x=XPos.LMARGIN)

    pdf.set_font("Times", size=16, style="B")
    pdf.cell(50, 8, text=f"Date nr.{invoice_date}", new_y=YPos.NEXT, new_x=XPos.LMARGIN)

    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    # Add a header row with column names
    columns = df.columns.tolist()
    columns = [item.replace("_", " ").title() for item in columns]
    pdf.set_font("Times", size=10, style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(30, 8, text=f"{columns[0]}", border=1, align="C")
    pdf.cell(70, 8, text=f"{columns[1]}", border=1, align="C")
    pdf.cell(30, 8, text=f"{columns[2]}", border=1, align="C")
    pdf.cell(30, 8, text=f"{columns[3]}", border=1, align="C")
    pdf.cell(30, 8, text=f"{columns[4]}", border=1, new_y=YPos.NEXT, new_x=XPos.LMARGIN, align="C")

    # Add data rows
    for index, row in df.iterrows():
        pdf.set_font("Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(30, 8, text=f"{row['product_id']}", border=1, align="C")
        pdf.cell(70, 8, text=f"{row['product_name']}", border=1, align="C")
        pdf.cell(30, 8, text=f"{row['amount_purchased']}", border=1, align="C")
        pdf.cell(30, 8, text=f"{row['price_per_unit']}", border=1, align="C")
        pdf.cell(30, 8, text=f"{row['total_price']}", border=1, new_y=YPos.NEXT, new_x=XPos.LMARGIN, align="C")

    # Add the total row
    total_sum = df['total_price'].sum()
    pdf.set_font("Times", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(30, 8, text="", border=1)
    pdf.cell(70, 8, text="", border=1)
    pdf.cell(30, 8, text="", border=1)
    pdf.cell(30, 8, text="", border=1)
    pdf.cell(30, 8, text=f"{total_sum}", border=1, new_y=YPos.NEXT, new_x=XPos.LMARGIN, align="C")

    # Add the total price text
    pdf.set_font("Times", size=10, style="B")
    pdf.cell(30, 8, text=f"The total price is: {total_sum}", new_y=YPos.NEXT, new_x=XPos.LMARGIN)

    # Add the Company and logo
    pdf.set_font("Times", size=14, style="B")
    pdf.cell(30, 8, text=f"Pythonhow")
    pdf.image("pythonhow.png", w=10)

    pdf.output(f"invoices/{filename}.pdf")